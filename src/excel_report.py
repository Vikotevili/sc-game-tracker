from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .config import EXCEL_PATH, GAME_END_DAY, MANAGEMENT_START_DAY


HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
TITLE_FONT = Font(bold=True, name="Calibri", size=16, color="1F4E79")
LABEL_FONT = Font(bold=True, name="Calibri", size=11)
GOOD_FILL = PatternFill("solid", fgColor="C6EFCE")
WARN_FILL = PatternFill("solid", fgColor="FFEB9C")
BAD_FILL = PatternFill("solid", fgColor="FFC7CE")
ALT_FILL = PatternFill("solid", fgColor="D6EAF8")
THIN = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

SNAPSHOT_HEADERS = [
    "抓取时间", "游戏日", "现金", "排名", "队伍数", "领先队伍", "距第一名差额",
    "当前产能", "计划产能", "运输方式", "订货点", "批量",
    "仓库库存", "在途", "WIP",
    "当前周期", "周期需求", "周期缺货", "周期交付", "周期服务水平", "周期现金变动",
]
SNAPSHOT_FORMATS = [
    None, "0", '"$"#,##0.00', "0", "0", None, '"$"#,##0.00',
    "0.00", "0.00", None, "#,##0", "#,##0",
    "#,##0.00", "#,##0.00", "#,##0.00",
    "0", "#,##0.00", "#,##0.00", "#,##0.00", "0.0%", '"$"#,##0.00',
]

PERIOD_HEADERS = [
    "周期", "起始日", "结束日", "天数", "是否完整", "接管后",
    "需求", "缺货", "交付", "服务水平", "出货",
    "期初现金", "期末现金", "现金变动",
    "期末仓库库存", "平均仓库库存", "期末在途", "期末WIP", "估算收入",
]
PERIOD_FORMATS = [
    "0", "0", "0", "0", None, None,
    "#,##0.00", "#,##0.00", "#,##0.00", "0.0%", "#,##0.00",
    '"$"#,##0.00', '"$"#,##0.00', '"$"#,##0.00',
    "#,##0.00", "#,##0.00", "#,##0.00", "#,##0.00", '"$"#,##0.00',
]

DAILY_HEADERS = [
    "游戏日", "接管后", "需求", "缺货", "交付", "服务水平", "出货",
    "现金", "仓库库存", "在途邮件", "在途卡车", "在途合计", "WIP", "估算收入",
]
DAILY_FORMATS = [
    "0", None, "#,##0.00", "#,##0.00", "#,##0.00", "0.0%", "#,##0.00",
    '"$"#,##0.00', "#,##0.00", "#,##0.00", "#,##0.00", "#,##0.00", "#,##0.00", '"$"#,##0.00',
]

CASH_HEADERS = ["抓取时间", "游戏日", "项目", "金额", "类型"]
CASH_FORMATS = [None, "0", None, '"$"#,##0.00', None]

PARAM_HEADERS = [
    "抓取时间", "游戏日", "工厂地区", "当前产能", "计划产能", "工厂运输",
    "订货点", "批量", "仓库运输", "是否运营",
]
PARAM_FORMATS = [None, "0", None, "0.00", "0.00", None, "#,##0", "#,##0", None, None]

HISTORY_HEADERS = ["游戏日", "参数", "参数(短)", "工厂", "仓库", "新值"]
HISTORY_FORMATS = ["0.00", None, None, None, None, "#,##0.00"]

STANDING_HEADERS = ["抓取时间", "游戏日", "排名", "队伍", "现金"]
STANDING_FORMATS = [None, "0", "0", None, '"$"#,##0.00']


def _style_header(ws: Worksheet, row: int, cols: int) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(row, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN
    ws.row_dimensions[row].height = 22
    ws.freeze_panes = f"A{row + 1}"
    ws.auto_filter.ref = f"A{row}:{get_column_letter(cols)}{row}"


def _autosize(ws: Worksheet, min_width: int = 10, max_width: int = 36) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = min_width
        for cell in col[:80]:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(max_width, len(value) + 2))
        ws.column_dimensions[letter].width = width


def _write_kv(ws: Worksheet, row: int, label: str, value: Any, num_format: str | None = None) -> int:
    ws.cell(row, 1, label).font = LABEL_FONT
    cell = ws.cell(row, 2, value)
    if num_format and isinstance(value, (int, float)):
        cell.number_format = num_format
    return row + 1


def _service_fill(value: float | None) -> PatternFill | None:
    if value is None:
        return None
    if value >= 0.95:
        return GOOD_FILL
    if value >= 0.80:
        return WARN_FILL
    return BAD_FILL


def _as_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _norm(value: Any) -> Any:
    number = _as_int(value)
    if number is not None and isinstance(value, (int, float)):
        return number
    return value


def _last_used_row(ws: Worksheet, col: int = 1, header_row: int = 1) -> int:
    last = header_row
    for row in range(header_row + 1, (ws.max_row or header_row) + 1):
        if ws.cell(row, col).value not in (None, ""):
            last = row
    return last


def _find_row(ws: Worksheet, col: int, key: Any, header_row: int = 1) -> int | None:
    want = _norm(key)
    for row in range(header_row + 1, (ws.max_row or header_row) + 1):
        if _norm(ws.cell(row, col).value) == want:
            return row
    return None


def _write_row(
    ws: Worksheet,
    row: int,
    values: list[Any],
    formats: list[str | None],
    headers: list[str],
) -> None:
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row, col, value)
        cell.border = THIN
        cell.alignment = Alignment(vertical="center")
        fmt = formats[col - 1] if col - 1 < len(formats) else None
        if fmt and isinstance(value, (int, float)):
            cell.number_format = fmt
        header = headers[col - 1] if col - 1 < len(headers) else ""
        if str(header).endswith("服务水平") and isinstance(value, float):
            fill = _service_fill(value)
            if fill:
                cell.fill = fill
        elif row % 2 == 0:
            cell.fill = ALT_FILL
    end = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{end}{max(row, 1)}"


def _ensure_sheet(wb, name: str, headers: list[str]) -> Worksheet:
    if name in wb.sheetnames:
        ws = wb[name]
        first = ws.cell(1, 1).value
        last = ws.cell(1, len(headers)).value
        if first == headers[0] and last == headers[-1]:
            return ws
        backup = f"{name}_旧版"
        if backup not in wb.sheetnames:
            ws.title = backup
    ws = wb.create_sheet(name)
    for idx, header in enumerate(headers, start=1):
        ws.cell(1, idx, header)
    _style_header(ws, 1, len(headers))
    return ws


def _period_values(item: dict[str, Any]) -> list[Any]:
    return [
        item["period"],
        item["start_day"],
        item["end_day"],
        item["days_in_period"],
        "是" if item["complete"] else "否",
        "是" if item["after_takeover"] else "否",
        item["demand"],
        item["lost_demand"],
        item["filled"],
        item["service_level"],
        item["shipments"],
        item["cash_start"],
        item["cash_end"],
        item["cash_change"],
        item["inventory_end"],
        item["inventory_avg"],
        item["pipeline_end"],
        item["wip_end"],
        item["est_revenue"],
    ]


def _daily_values(item: dict[str, Any]) -> list[Any]:
    return [
        item["day"],
        "是" if item["after_takeover"] else "否",
        item["demand"],
        item["lost_demand"],
        item["filled"],
        item["service_level"],
        item["shipments"],
        item["cash"],
        item["inventory"],
        item["pipeline_mail"],
        item["pipeline_truck"],
        item["pipeline"],
        item["wip"],
        item["est_revenue"],
    ]


def _write_overview(wb, report: dict[str, Any]) -> None:
    if "概览" in wb.sheetnames:
        index = wb.sheetnames.index("概览")
        del wb["概览"]
        ws = wb.create_sheet("概览", index)
    else:
        ws = wb.create_sheet("概览", 0)

    header = report["header"]
    rank = report["rank"]
    factory = report["factory"]
    warehouse = report["warehouse"]
    stock = report["stock"]
    current = report["current_period"] or {}
    last_complete = report["last_complete_period"] or {}

    ws["A1"] = "Supply Chain Game 运营看板（最新）"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")
    ws["A2"] = "本页只显示最新状态。历史抓取、每日和三日周期都是追加，不会覆盖已有行。"
    ws.merge_cells("A2:D2")

    row = 4
    row = _write_kv(ws, row, "抓取时间", report["fetched_at"])
    row = _write_kv(ws, row, "团队", header.get("team"))
    row = _write_kv(ws, row, "当前游戏日", report["day"], "#,##0")
    row = _write_kv(ws, row, "剩余游戏日", report["days_left"], "#,##0")
    row = _write_kv(ws, row, "游戏结束日", GAME_END_DAY, "#,##0")
    row = _write_kv(ws, row, "接管起始日", MANAGEMENT_START_DAY, "#,##0")
    row = _write_kv(ws, row, "当前现金", header.get("cash"), '"$"#,##0.00')
    row = _write_kv(ws, row, "当前排名", rank.get("rank"), "0")
    row = _write_kv(ws, row, "参赛队伍数", rank.get("teams"), "0")
    row = _write_kv(ws, row, "领先队伍", rank.get("leader_team"))
    row = _write_kv(ws, row, "距第一名差额", rank.get("gap_to_leader"), '"$"#,##0.00')

    row += 1
    ws.cell(row, 1, "工厂 / 仓库参数").font = TITLE_FONT
    row += 1
    row = _write_kv(ws, row, "工厂地区", factory.get("region_name"))
    row = _write_kv(ws, row, "当前产能 (鼓/日)", factory.get("current_capacity"), "0.00")
    row = _write_kv(ws, row, "计划产能 (鼓/日)", factory.get("scheduled_capacity"), "0.00")
    row = _write_kv(ws, row, "工厂运输方式", factory.get("shipping"))
    row = _write_kv(ws, row, "订货点 ROP", factory.get("order_point"), "#,##0")
    row = _write_kv(ws, row, "订货批量 Q", factory.get("order_quantity"), "#,##0")
    row = _write_kv(ws, row, "仓库运输方式", warehouse.get("shipping"))
    row = _write_kv(ws, row, "期末仓库库存", stock.get("inventory"), "#,##0.00")
    row = _write_kv(ws, row, "期末在途库存", stock.get("pipeline"), "#,##0.00")
    row = _write_kv(ws, row, "期末在制品 WIP", stock.get("wip"), "#,##0.00")

    row += 1
    ws.cell(row, 1, "当前 3 日周期").font = TITLE_FONT
    row += 1
    row = _write_kv(ws, row, "周期编号", current.get("period"))
    row = _write_kv(ws, row, "覆盖游戏日", f"{current.get('start_day')}–{current.get('end_day')}" if current else None)
    row = _write_kv(ws, row, "是否完整周期", "是" if current.get("complete") else "否（进行中）")
    row = _write_kv(ws, row, "周期需求", current.get("demand"), "#,##0.00")
    row = _write_kv(ws, row, "周期缺货", current.get("lost_demand"), "#,##0.00")
    row = _write_kv(ws, row, "周期交付", current.get("filled"), "#,##0.00")
    row = _write_kv(ws, row, "周期服务水平", current.get("service_level"), "0.0%")
    row = _write_kv(ws, row, "周期出货", current.get("shipments"), "#,##0.00")
    row = _write_kv(ws, row, "周期现金变动", current.get("cash_change"), '"$"#,##0.00')

    row += 1
    ws.cell(row, 1, "上一完整 3 日周期").font = TITLE_FONT
    row += 1
    row = _write_kv(ws, row, "周期编号", last_complete.get("period"))
    row = _write_kv(
        ws,
        row,
        "覆盖游戏日",
        f"{last_complete.get('start_day')}–{last_complete.get('end_day')}" if last_complete else None,
    )
    row = _write_kv(
        ws,
        row,
        "需求 / 缺货 / 交付",
        last_complete and f"{last_complete.get('demand')} / {last_complete.get('lost_demand')} / {last_complete.get('filled')}",
    )
    row = _write_kv(ws, row, "服务水平", last_complete.get("service_level"), "0.0%")
    row = _write_kv(ws, row, "现金变动", last_complete.get("cash_change"), '"$"#,##0.00')

    row += 2
    ws.cell(row, 1, "说明").font = LABEL_FONT
    ws.cell(row + 1, 1, "时间序列请看「快照历史」「每日数据」「三日周期」。每次抓取只追加新行；已完成的历史行不会被改写。")
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=4)
    _autosize(ws, 16, 42)
    ws.column_dimensions["B"].width = 28


def _append_snapshot(ws: Worksheet, report: dict[str, Any]) -> None:
    header = report["header"]
    rank = report["rank"]
    factory = report["factory"]
    stock = report["stock"]
    current = report["current_period"] or {}
    last = _last_used_row(ws)
    if last > 1 and ws.cell(last, 1).value == report["fetched_at"]:
        return
    if last > 1:
        same_day = _as_int(ws.cell(last, 2).value) == report["day"]
        same_cash = ws.cell(last, 3).value == header.get("cash")
        same_rank = _as_int(ws.cell(last, 4).value) == rank.get("rank")
        same_inv = ws.cell(last, 13).value == stock.get("inventory")
        if same_day and same_cash and same_rank and same_inv:
            return
    values = [
        report["fetched_at"],
        report["day"],
        header.get("cash"),
        rank.get("rank"),
        rank.get("teams"),
        rank.get("leader_team"),
        rank.get("gap_to_leader"),
        factory.get("current_capacity"),
        factory.get("scheduled_capacity"),
        factory.get("shipping"),
        factory.get("order_point"),
        factory.get("order_quantity"),
        stock.get("inventory"),
        stock.get("pipeline"),
        stock.get("wip"),
        current.get("period"),
        current.get("demand"),
        current.get("lost_demand"),
        current.get("filled"),
        current.get("service_level"),
        current.get("cash_change"),
    ]
    _write_row(ws, last + 1, values, SNAPSHOT_FORMATS, SNAPSHOT_HEADERS)


def _upsert_periods(ws: Worksheet, report: dict[str, Any]) -> None:
    current_day = report["day"]
    for item in report["periods"]:
        row = _find_row(ws, 1, item["period"])
        if row is None:
            _write_row(ws, _last_used_row(ws) + 1, _period_values(item), PERIOD_FORMATS, PERIOD_HEADERS)
            continue
        complete = ws.cell(row, 5).value == "是"
        if complete:
            continue
        if item["complete"] or item["end_day"] >= current_day or item["period"] == (report["current_period"] or {}).get("period"):
            _write_row(ws, row, _period_values(item), PERIOD_FORMATS, PERIOD_HEADERS)


def _upsert_daily(ws: Worksheet, report: dict[str, Any]) -> None:
    current_day = report["day"]
    for item in report["daily"]:
        row = _find_row(ws, 1, item["day"])
        if row is None:
            _write_row(ws, _last_used_row(ws) + 1, _daily_values(item), DAILY_FORMATS, DAILY_HEADERS)
        elif item["day"] == current_day:
            _write_row(ws, row, _daily_values(item), DAILY_FORMATS, DAILY_HEADERS)


def _append_cash(ws: Worksheet, report: dict[str, Any]) -> None:
    last = _last_used_row(ws)
    if last > 1 and ws.cell(last, 1).value == report["fetched_at"]:
        return
    for item in report["cash_status"]:
        last += 1
        _write_row(
            ws,
            last,
            [report["fetched_at"], report["day"], item["description"], item["amount"], item["kind"]],
            CASH_FORMATS,
            CASH_HEADERS,
        )


def _append_params(ws: Worksheet, report: dict[str, Any]) -> None:
    factory = report["factory"]
    warehouse = report["warehouse"]
    last = _last_used_row(ws)
    values = [
        report["fetched_at"],
        report["day"],
        factory.get("region_name"),
        factory.get("current_capacity"),
        factory.get("scheduled_capacity"),
        factory.get("shipping"),
        factory.get("order_point"),
        factory.get("order_quantity"),
        warehouse.get("shipping"),
        "是" if factory.get("operational") else "否",
    ]
    if last > 1:
        prev = [ws.cell(last, col).value for col in range(3, 11)]
        if prev == values[2:]:
            return
    _write_row(ws, last + 1, values, PARAM_FORMATS, PARAM_HEADERS)


def _append_history(ws: Worksheet, report: dict[str, Any]) -> None:
    seen: set[tuple[Any, ...]] = set()
    for row in range(2, _last_used_row(ws) + 1):
        seen.add(tuple(_norm(ws.cell(row, col).value) for col in range(1, 7)))
    last = _last_used_row(ws)
    for item in report["history"]:
        values = [
            item.get("day"),
            item.get("parameter"),
            item.get("parameter_short"),
            item.get("factory"),
            item.get("warehouse"),
            item.get("new_value"),
        ]
        key = tuple(_norm(v) for v in values)
        if key in seen:
            continue
        last += 1
        seen.add(key)
        _write_row(ws, last, values, HISTORY_FORMATS, HISTORY_HEADERS)


def _append_standing(ws: Worksheet, report: dict[str, Any]) -> None:
    last = _last_used_row(ws)
    if last > 1 and ws.cell(last, 1).value == report["fetched_at"]:
        return
    my_team = report["header"].get("team")
    for item in report["standing"]:
        last += 1
        _write_row(
            ws,
            last,
            [report["fetched_at"], report["day"], item["rank"], item["team"], item["cash"]],
            STANDING_FORMATS,
            STANDING_HEADERS,
        )
        if item["team"] == my_team:
            for col in range(1, 6):
                ws.cell(last, col).fill = WARN_FILL


def _open_workbook(path: Path):
    if path.exists():
        return load_workbook(path)
    wb = Workbook()
    default = wb.active
    default.title = "概览"
    return wb


def _save_workbook(wb, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(path)
        return path
    except PermissionError:
        fallback = path.with_name(path.stem + "_pending.xlsx")
        wb.save(fallback)
        return fallback


def write_excel(report: dict[str, Any], path: Path = EXCEL_PATH) -> Path:
    wb = _open_workbook(path)
    _write_overview(wb, report)

    snapshot = _ensure_sheet(wb, "快照历史", SNAPSHOT_HEADERS)
    _append_snapshot(snapshot, report)
    _autosize(snapshot, 12, 22)

    periods = _ensure_sheet(wb, "三日周期", PERIOD_HEADERS)
    _upsert_periods(periods, report)
    _autosize(periods, 10, 16)

    daily = _ensure_sheet(wb, "每日数据", DAILY_HEADERS)
    _upsert_daily(daily, report)
    _autosize(daily, 10, 14)

    cash = _ensure_sheet(wb, "资金构成", CASH_HEADERS)
    _append_cash(cash, report)
    _autosize(cash, 14, 36)

    params = _ensure_sheet(wb, "运营参数", PARAM_HEADERS)
    _append_params(params, report)
    _autosize(params, 12, 18)

    history = _ensure_sheet(wb, "决策历史", HISTORY_HEADERS)
    _append_history(history, report)
    _autosize(history, 12, 48)

    standing = _ensure_sheet(wb, "排行榜", STANDING_HEADERS)
    _append_standing(standing, report)
    _autosize(standing, 12, 28)

    return _save_workbook(wb, path)
